"""
Полный бэкап базы в один .xlsx — по листу на каждую сущность, с вычисляемыми
полями (баланс клиента, прибыль/оплата заказа и т.п.), чтобы снимок отражал
всю логику системы (CLAUDE.md §4).
"""
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse


def _cell(v):
    """Привести значение к типу, понятному Excel."""
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (datetime, date)):
        return v
    if v is None:
        return ""
    return v


def _add_sheet(wb, title, header, rows):
    from openpyxl.styles import Font

    ws = wb.create_sheet(title[:31])  # Excel ограничивает длину имени листа
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append([_cell(v) for v in row])
    # авто-ширина колонок (грубо, по заголовкам/первым строкам)
    for i, _ in enumerate(header, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[letter].width = max(12, min(40, len(str(header[i - 1])) + 4))
    return ws


def build_backup_workbook():
    from openpyxl import Workbook

    from apps.clients.models import BalanceMovement, Client
    from apps.clients.services import (
        client_balance, client_deposits, client_due, client_refunds,
    )
    from apps.orders.models import Order, OrderExpense, OrderItem, Return
    from apps.orders.services import (
        item_returned_qty, item_sold_qty, order_calculation, order_paid,
        order_status,
    )
    from apps.warehouse.models import WarehouseItem
    from apps.expenses.models import Expense, ExpenseCategory, RecurringExpense
    from apps.finance.models import Investment, Investor, Reserve, ReserveMovement
    from apps.finance.services import investments_pool, reserve_balance
    from apps.accounts.roles import user_role
    from django.contrib.auth.models import Group, User

    wb = Workbook()
    wb.remove(wb.active)  # убрать пустой лист по умолчанию

    # --- Клиенты ---
    _add_sheet(wb, "Клиенты",
        ["ID", "ФИО", "Телефон", "Дата рождения", "Архив", "Создан",
         "Внесено", "Возвраты", "К оплате", "Баланс", "Заметки"],
        [[c.id, c.full_name, c.phone, c.birth_date, c.is_archived,
          c.created_at.replace(tzinfo=None),
          client_deposits(c), client_refunds(c), client_due(c), client_balance(c),
          c.notes]
         for c in Client.objects.all()])

    # --- Движения баланса ---
    _add_sheet(wb, "Движения баланса",
        ["ID", "Клиент", "Заказ", "Направление", "Сумма", "Способ",
         "Комментарий", "Дата", "Создано"],
        [[m.id, m.client.full_name, m.order_id, m.get_direction_display(),
          m.amount, m.get_method_display(), m.comment, m.paid_at,
          m.created_at.replace(tzinfo=None)]
         for m in BalanceMovement.objects.select_related("client")])

    # --- Заказы ---
    order_rows = []
    for o in Order.objects.select_related("client", "created_by"):
        c = order_calculation(o)
        st = order_status(o)
        order_rows.append([
            o.id, o.client.full_name,
            o.created_by.username if o.created_by else "",
            o.is_archived, o.created_at.replace(tzinfo=None), st["label"],
            c["revenue"], c["cost"], c["extra_expenses"], c["profit"],
            c["delivery"], c["due"], order_paid(o),
            c["due"] - order_paid(o), o.notes,
        ])
    _add_sheet(wb, "Заказы",
        ["ID", "Клиент", "Оформил", "Архив", "Создан", "Статус", "Выручка",
         "Себестоимость", "Доп.расходы", "Прибыль", "Доставка", "К оплате",
         "Оплачено", "Остаток", "Заметки"],
        order_rows)

    # --- Позиции заказов ---
    _add_sheet(wb, "Позиции заказов",
        ["ID", "Заказ", "Наименование", "Со склада (ID)", "Себес валюта",
         "Валюта", "Себес ₸", "Кол-во", "Выдано", "Продано", "Возвращено",
         "Цена продажи", "Доставка/ед", "Страна", "Сайт", "Трек", "Статус",
         "Дата покупки", "Дата доставки"],
        [[i.id, i.order_id, i.name, i.warehouse_item_id, i.cost_foreign,
          i.currency, i.cost_kzt, i.qty, i.issued_qty, item_sold_qty(i),
          item_returned_qty(i), i.sale_price, i.delivery_price, i.country,
          i.site, i.track_number, i.get_status_display(),
          i.purchase_date, i.delivery_date]
         for i in OrderItem.objects.all()])

    # --- Возвраты ---
    _add_sheet(wb, "Возвраты",
        ["ID", "Позиция (ID)", "Кол-во", "Что с товаром", "Возврат клиенту",
         "Дата", "Комментарий"],
        [[r.id, r.order_item_id, r.qty, r.get_disposition_display(),
          r.refund_amount, r.return_date, r.comment]
         for r in Return.objects.all()])

    # --- Доп.расходы заказа ---
    _add_sheet(wb, "Доп.расходы заказа",
        ["ID", "Заказ", "Тип", "Сумма", "Комментарий"],
        [[e.id, e.order_id, e.get_type_display(), e.amount, e.comment]
         for e in OrderExpense.objects.all()])

    # --- Склад ---
    _add_sheet(wb, "Склад",
        ["ID", "Наименование", "Страна", "Себес валюта", "Валюта", "Себес ₸",
         "Кол-во", "Доставка", "Прочее", "Плановая цена", "Полная затрата",
         "Статус", "Архив", "Дата покупки", "Дата доставки", "Заметки"],
        [[w.id, w.name, w.country, w.cost_foreign, w.currency, w.cost_kzt,
          w.qty, w.delivery_cost, w.other_costs, w.planned_price, w.full_cost,
          w.get_status_display(), w.is_archived, w.purchase_date,
          w.delivery_date, w.notes]
         for w in WarehouseItem.objects.all()])

    # --- Категории расходов ---
    _add_sheet(wb, "Категории расходов",
        ["ID", "Название", "Ежемесячная"],
        [[c.id, c.name, c.is_recurring] for c in ExpenseCategory.objects.all()])

    # --- Ежемесячные напоминания ---
    _add_sheet(wb, "Ежемес. напоминания",
        ["ID", "Название", "Категория", "Ориентир", "Способ", "Активно", "Заметка"],
        [[r.id, r.name, r.category.name, r.planned_amount,
          r.get_method_display(), r.is_active, r.notes]
         for r in RecurringExpense.objects.select_related("category")])

    # --- Расходы (факт) ---
    _add_sheet(wb, "Расходы",
        ["ID", "Категория", "Сумма", "Способ", "Дата выдачи", "За месяц",
         "Гасит напоминание", "Комментарий"],
        [[e.id, e.category.name, e.amount, e.get_method_display(),
          e.expense_date, e.period,
          e.recurring.name if e.recurring else "", e.comment]
         for e in Expense.objects.select_related("category", "recurring")])

    # --- Инвесторы / Инвестиции ---
    _add_sheet(wb, "Инвесторы",
        ["ID", "Имя", "Заметки"],
        [[i.id, i.name, i.notes] for i in Investor.objects.all()])
    _add_sheet(wb, "Инвестиции",
        ["ID", "Инвестор", "Направление", "Сумма", "Способ", "Дата", "Комментарий"],
        [[i.id, i.investor.name, i.get_direction_display(), i.amount,
          i.get_method_display(), i.moved_at, i.comment]
         for i in Investment.objects.select_related("investor")])

    # --- Резервы / движения ---
    _add_sheet(wb, "Резервы",
        ["ID", "Название", "Тип", "Цель", "Баланс", "Комментарий"],
        [[r.id, r.name, r.get_kind_display(), r.target_amount,
          reserve_balance(r), r.comment]
         for r in Reserve.objects.all()])
    _add_sheet(wb, "Движения резервов",
        ["ID", "Резерв", "Направление", "Сумма", "Дата", "Комментарий"],
        [[m.id, m.reserve.name, m.get_direction_display(), m.amount,
          m.moved_at, m.comment]
         for m in ReserveMovement.objects.select_related("reserve")])

    # --- Пользователи / Роли ---
    _add_sheet(wb, "Пользователи",
        ["ID", "Логин", "Имя", "Фамилия", "Email", "Активен", "Суперпользователь",
         "Роль", "Роли (группы)"],
        [[u.id, u.username, u.first_name, u.last_name, u.email, u.is_active,
          u.is_superuser, user_role(u) or "",
          ", ".join(g.name for g in u.groups.all())]
         for u in User.objects.all()])
    _add_sheet(wb, "Роли",
        ["ID", "Название", "Пользователей", "Прав"],
        [[g.id, g.name, g.user_set.count(), g.permissions.count()]
         for g in Group.objects.all()])

    return wb


def backup_excel_response() -> HttpResponse:
    wb = build_backup_workbook()
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = 'attachment; filename="order-ledger-backup.xlsx"'
    return response
